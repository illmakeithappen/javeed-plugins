"""Render plan and input data to multi-sheet XLSX workbooks."""

from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path
from typing import Any

from .schemas import (
    ABSENCES_COLS,
    ASSIGNMENTS_COLS,
    DIRECTIVE_COMPLIANCE_COLS,
    DIRECTIVES_COLS,
    EMPLOYEES_COLS,
    EVAL_MATRIX_COLS,
    FAIRNESS_COLS,
    OPEN_SLOTS_COLS,
    SCORE_COMPONENTS,
    SHIFTS_COLS,
    UNASSIGNED_COLS,
    fmt_bool,
    pipe_join,
)

# assignment_kind -> short label for CSV
_KIND_MAP = {
    "applicant": "applicant",
    "recommendation_without_applicant": "recommendation_without",
    "recommendation_despite_applicants": "recommendation_despite",
}

_WEEKDAY_DE = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

# Employee Rules columns for the flattened profile view
_EMPLOYEE_RULES_COLS = [
    "employee_name",
    "target_weekly_hours",
    "max_monthly_hours",
    "max_weekly_hours",
    "no_additional_shifts",
    "preferred_working_areas",
    "preferred_shift_types",
    "notes",
    "source",
]

# Rule fields that we track sources for
_RULE_FIELDS = [
    "target_weekly_hours",
    "max_monthly_hours",
    "max_weekly_hours",
    "no_additional_shifts",
    "preferred_working_areas",
    "preferred_shift_types",
    "notes",
]

# Employee Summary columns (merged employee + profile + fairness)
_EMPLOYEE_SUMMARY_COLS = [
    "employee_name",
    "employee_id",
    "role",
    "employment",
    "hourly_wage",
    "max_salary",
    "skills",
    "target_weekly_hours",
    "notes",
    "rule_source",
    "assigned_hours",
    "assigned_slots",
    "delta_to_target",
]

# Review Data columns
_REVIEW_DATA_COLS = [
    "date",
    "weekday",
    "start",
    "end",
    "shift_type",
    "area",
    "employee_name",
    "is_applicant",
    "score",
    "reasoning",
    "alternative_1",
    "alternative_2",
    "alternative_3",
]


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _rule_source_summary(name: str, rule_sources: dict[str, dict[str, str]]) -> str:
    """Summarise where an employee's rules came from.

    Returns one of: 'ordio', 'vorgaben', 'ordio+vorgaben', or '' (no rules).
    """
    sources_for_name = rule_sources.get(name, {})
    if not sources_for_name:
        return ""
    unique = set(sources_for_name.values())
    if unique == {"ordio"}:
        return "ordio"
    if unique == {"vorgaben"}:
        return "vorgaben"
    return "ordio+vorgaben"


def _get_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        return Workbook, Font, PatternFill
    except ImportError as exc:
        raise ImportError("openpyxl is required for XLSX export: pip install openpyxl") from exc


def _style_headers(worksheets):
    """Apply bold + blue fill to header row of each worksheet."""
    _, Font, PatternFill = _get_openpyxl()
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for ws in worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill


def _read_csv_file(path: Path) -> list[dict[str, str]]:
    """Read a CSV file into list of dicts."""
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _read_json_file(path: Path) -> dict:
    """Read a JSON file."""
    if not path.exists():
        return {}
    with open(path, encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Row builders for output XLSX (existing)
# ---------------------------------------------------------------------------

def _assignment_row(a: dict[str, Any]) -> dict[str, Any]:
    sd = a.get("score_detail", {})
    return {
        "slot_id": a.get("slot_id", ""),
        "date": a.get("date", ""),
        "start": a.get("start", ""),
        "end": a.get("end", ""),
        "shift_type": a.get("shift_type", ""),
        "area": a.get("working_area", ""),
        "employee_id": a.get("ordio_employee_id", ""),
        "employee_name": a.get("employee_name", ""),
        "score": a.get("score", 0),
        "is_applicant": fmt_bool(a.get("is_applicant", False)),
        "kind": _KIND_MAP.get(a.get("assignment_kind", ""), a.get("assignment_kind", "")),
        **{f"score_{c}": sd.get(c, 0) for c in SCORE_COMPONENTS},
        "reasons": pipe_join(a.get("reasons", [])),
        "target_hours": a.get("target_hours") if a.get("target_hours") is not None else "",
        "month_hours_before": a.get("run_month_hours_before", 0),
    }


def _unassigned_row(u: dict[str, Any]) -> dict[str, Any]:
    top = u.get("top_candidates", [])
    blocked_parts = []
    for c in top:
        reasons = c.get("blocked_reasons", [])
        if reasons:
            blocked_parts.append(f"{c.get('employee_name', '')}:{','.join(reasons)}")
    return {
        "slot_id": u.get("slot_id", ""),
        "date": u.get("date", ""),
        "start": u.get("start", ""),
        "end": u.get("end", ""),
        "shift_type": u.get("shift_type", ""),
        "area": u.get("working_area", ""),
        "reason": u.get("reason", ""),
        "top_blocked": "|".join(blocked_parts),
    }


def _fairness_row(f: dict[str, Any]) -> dict[str, Any]:
    return {
        "employee_name": f.get("employee_name", ""),
        "assigned_hours": f.get("assigned_hours", 0),
        "assigned_slots": f.get("assigned_slots", 0),
        "target_hours": f.get("target_hours") if f.get("target_hours") is not None else "",
        "delta": f.get("delta_to_target") if f.get("delta_to_target") is not None else "",
    }


# ---------------------------------------------------------------------------
# German reasoning builder (Begruendungs-Bausteine)
# ---------------------------------------------------------------------------

def _build_reasoning(a: dict[str, Any]) -> str:
    """Build German-language reasoning text from an assignment's score_detail."""
    sd = a.get("score_detail", {})
    parts: list[str] = []

    if a.get("is_applicant"):
        parts.append("Bewerbung eingegangen. Bewerber-Bonus: +80.")

    target = a.get("target_hours")
    rest_score = sd.get("rest", 0)
    if rest_score and target:
        month_before = a.get("run_month_hours_before", 0)
        remaining = round(target - month_before, 1)
        parts.append(f"Restkapazitaet: {remaining}h von {target}h Ziel.")

    fairness_score = sd.get("fairness", 0)
    if fairness_score:
        parts.append(f"Fairness: {fairness_score}/30.")

    skill_score = sd.get("skill", 0)
    if skill_score:
        area = a.get("working_area", "")
        parts.append(f"Skills: {area}-Match ({skill_score}).")

    role_score = sd.get("role", 0)
    if role_score:
        parts.append(f"Rollenfit: {role_score}.")

    fixed_score = sd.get("fixed", 0)
    if fixed_score:
        parts.append("Regelmaessige Schicht (historisches Muster).")

    pref_score = sd.get("preference", 0)
    if pref_score:
        parts.append("Passt zu Schichtpraeferenz.")

    salary_score = sd.get("salary", 0)
    if salary_score and salary_score < 0:
        parts.append("Hinweis: Nahe an Gehaltsgrenze (-12).")

    parts.append("Geprueft: Ruhezeit OK, Monatsgrenze OK.")

    score = a.get("score", 0)
    parts.append(f"Gesamt: {score}.")

    return " ".join(parts)


def _weekday_de(iso_date: str) -> str:
    """Convert ISO date string to German weekday abbreviation."""
    try:
        d = date.fromisoformat(iso_date)
        return _WEEKDAY_DE[d.weekday()]
    except (ValueError, TypeError):
        return ""


def _format_date_de(iso_date: str) -> str:
    """Convert YYYY-MM-DD to DD.MM. format."""
    try:
        d = date.fromisoformat(iso_date)
        return f"{d.day:02d}.{d.month:02d}."
    except (ValueError, TypeError):
        return iso_date


# ---------------------------------------------------------------------------
# Workbook 1: Input XLSX
# ---------------------------------------------------------------------------

def render_input_xlsx(directory: Path, path: Path) -> Path:
    """Render CSV input directory to a multi-sheet XLSX workbook.

    Sheets: Overview, Employees, Open Slots, Assigned Shifts, Absences, Employee Rules.
    """
    Workbook, Font, PatternFill = _get_openpyxl()
    directory = Path(directory)
    path = Path(path)

    meta = _read_json_file(directory / "meta.json")
    profile = _read_json_file(directory / "profile.json")
    employees = _read_csv_file(directory / "employees.csv")
    shifts = _read_csv_file(directory / "shifts.csv")
    open_slots = _read_csv_file(directory / "open_slots.csv")
    absences = _read_csv_file(directory / "absences.csv")

    wb = Workbook()

    # --- Overview sheet ---
    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.append(["Field", "Value"])
    overview_fields = [
        ("betrieb", meta.get("betrieb", "")),
        ("snapshot_id", meta.get("snapshot_id", "")),
        ("range_from", meta.get("range_from", "")),
        ("range_to", meta.get("range_to", "")),
        ("profile", profile.get("name", "")),
        ("employees", len(employees)),
        ("assigned_shifts", len(shifts)),
        ("open_slots", len(open_slots)),
        ("absences", len(absences)),
        ("employee_rules", len(profile.get("employee_rules", {}))),
    ]
    for field, value in overview_fields:
        ws_overview.append([field, value])

    # --- Employees sheet ---
    ws_emp = wb.create_sheet("Employees")
    ws_emp.append(EMPLOYEES_COLS)
    for row in employees:
        ws_emp.append([row.get(c, "") for c in EMPLOYEES_COLS])

    # --- Open Slots sheet ---
    ws_slots = wb.create_sheet("Open Slots")
    ws_slots.append(OPEN_SLOTS_COLS)
    for row in open_slots:
        ws_slots.append([row.get(c, "") for c in OPEN_SLOTS_COLS])

    # --- Assigned Shifts sheet ---
    ws_shifts = wb.create_sheet("Assigned Shifts")
    ws_shifts.append(SHIFTS_COLS)
    for row in shifts:
        ws_shifts.append([row.get(c, "") for c in SHIFTS_COLS])

    # --- Absences sheet ---
    ws_abs = wb.create_sheet("Absences")
    ws_abs.append(ABSENCES_COLS)
    for row in absences:
        ws_abs.append([row.get(c, "") for c in ABSENCES_COLS])

    # --- Rule sources (ordio vs vorgaben) ---
    # Use pre-computed _rule_sources from profile.json (written by ingest step).
    # Only fall back to tagging everything as "ordio" when no sources are stored.
    rule_sources: dict[str, dict[str, str]] = profile.get("_rule_sources", {})
    if not rule_sources:
        rule_sources = {
            name: {field: "ordio" for field in rules}
            for name, rules in profile.get("employee_rules", {}).items()
        }

    # --- Employee Rules sheet (flattened, with source) ---
    ws_rules = wb.create_sheet("Employee Rules")
    ws_rules.append(_EMPLOYEE_RULES_COLS)
    for name, rules in profile.get("employee_rules", {}).items():
        ws_rules.append([
            name,
            rules.get("target_weekly_hours", ""),
            rules.get("max_monthly_hours", ""),
            rules.get("max_weekly_hours", ""),
            rules.get("no_additional_shifts", ""),
            pipe_join(rules.get("preferred_working_areas", [])),
            pipe_join(rules.get("preferred_shift_types", [])),
            rules.get("notes", ""),
            _rule_source_summary(name, rule_sources),
        ])

    all_sheets = [ws_overview, ws_emp, ws_slots, ws_shifts, ws_abs, ws_rules]

    # --- Vorgaben / Directives sheet ---
    vorgaben_path = directory / "vorgaben.txt"
    directives_csv = _read_csv_file(directory / "directives.csv")

    if vorgaben_path.exists() and vorgaben_path.read_text(encoding="utf-8").strip():
        ws_vor = wb.create_sheet("Vorgaben")
        ws_vor.append(["text"])
        for line in vorgaben_path.read_text(encoding="utf-8").strip().splitlines():
            ws_vor.append([line])
        all_sheets.append(ws_vor)
    elif directives_csv:
        ws_dir = wb.create_sheet("Directives")
        ws_dir.append(DIRECTIVES_COLS)
        for row in directives_csv:
            ws_dir.append([row.get(c, "") for c in DIRECTIVES_COLS])
        all_sheets.append(ws_dir)

    _style_headers(all_sheets)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# Workbook 2: Output XLSX (enhanced)
# ---------------------------------------------------------------------------

def render_xlsx(
    plan: dict[str, Any],
    path: Path,
    *,
    snapshot: dict[str, Any] | None = None,
    profile: dict[str, Any] | None = None,
    golden_comparison: dict[str, Any] | None = None,
) -> Path:
    """Render a plan to a multi-sheet XLSX workbook.

    Base sheets (always): Assignments, Unassigned, Fairness, Metrics.
    Enhanced sheets (when snapshot + profile provided): Employee Summary, Guide, Review Data.

    Returns the path to the written file.
    """
    Workbook, Font, PatternFill = _get_openpyxl()

    wb = Workbook()
    all_sheets = []

    # --- Assignments sheet ---
    ws_assign = wb.active
    ws_assign.title = "Assignments"
    ws_assign.append(ASSIGNMENTS_COLS)
    for a in plan.get("assignments", []):
        row = _assignment_row(a)
        ws_assign.append([row.get(c, "") for c in ASSIGNMENTS_COLS])
    all_sheets.append(ws_assign)

    # --- Unassigned sheet ---
    ws_unassigned = wb.create_sheet("Unassigned")
    ws_unassigned.append(UNASSIGNED_COLS)
    for u in plan.get("unassigned", []):
        row = _unassigned_row(u)
        ws_unassigned.append([row.get(c, "") for c in UNASSIGNED_COLS])
    all_sheets.append(ws_unassigned)

    # --- Fairness sheet ---
    ws_fairness = wb.create_sheet("Fairness")
    ws_fairness.append(FAIRNESS_COLS)
    for f in plan.get("fairness", []):
        row = _fairness_row(f)
        ws_fairness.append([row.get(c, "") for c in FAIRNESS_COLS])
    all_sheets.append(ws_fairness)

    # --- Metrics sheet ---
    ws_metrics = wb.create_sheet("Metrics")
    metrics = plan.get("metrics", {})
    meta_fields = [
        ("plan_id", plan.get("plan_id", "")),
        ("generated_at", plan.get("generated_at", "")),
        ("snapshot_id", plan.get("snapshot_id", "")),
        ("betrieb", plan.get("betrieb", "")),
        ("range_from", plan.get("range", {}).get("from", "")),
        ("range_to", plan.get("range", {}).get("to", "")),
        ("profile", plan.get("profile", "")),
        ("total_slots", metrics.get("total_slots", 0)),
        ("assigned", metrics.get("assigned_slots", 0)),
        ("unassigned", metrics.get("unassigned_slots", 0)),
        ("fill_rate", metrics.get("fill_rate", 0)),
    ]
    ws_metrics.append(["Field", "Value"])
    for field_name, value in meta_fields:
        ws_metrics.append([field_name, value])
    all_sheets.append(ws_metrics)

    # --- Enhanced sheets (only when snapshot + profile provided) ---
    if snapshot is not None and profile is not None:
        # Employee Summary
        ws_summary = wb.create_sheet("Employee Summary")
        ws_summary.append(_EMPLOYEE_SUMMARY_COLS)

        # Build fairness lookup: employee_name -> fairness row
        fairness_lookup: dict[str, dict] = {}
        for f in plan.get("fairness", []):
            fairness_lookup[f.get("employee_name", "")] = f

        employee_rules = profile.get("employee_rules", {})
        rule_sources = profile.get("_rule_sources", {})
        for emp in snapshot.get("employees", []):
            name = emp.get("full_name", "")
            rules = employee_rules.get(name, {})
            fair = fairness_lookup.get(name, {})
            ws_summary.append([
                name,
                emp.get("ordio_employee_id", ""),
                emp.get("role", ""),
                emp.get("employment", ""),
                emp.get("hourly_wage", ""),
                emp.get("max_salary", ""),
                pipe_join(emp.get("skills", [])),
                rules.get("target_weekly_hours", ""),
                rules.get("notes", ""),
                _rule_source_summary(name, rule_sources),
                fair.get("assigned_hours", 0),
                fair.get("assigned_slots", 0),
                fair.get("delta_to_target") if fair.get("delta_to_target") is not None else "",
            ])
        all_sheets.append(ws_summary)

        # Guide sheet (static reference)
        ws_guide = wb.create_sheet("Guide")
        _write_guide_sheet(ws_guide)
        all_sheets.append(ws_guide)

        # Freigabe sheet (interactive review/approval mask)
        ws_freigabe = wb.create_sheet("Freigabe")
        _write_freigabe_sheet(ws_freigabe, plan)
        # Freigabe is not added to all_sheets (has its own header styling)

        # Review Data sheet (raw reference)
        ws_review = wb.create_sheet("Review Data")
        ws_review.append(_REVIEW_DATA_COLS)
        for a in plan.get("assignments", []):
            reasoning = _build_reasoning(a)
            alts = a.get("alternatives", [])
            alt_strs = []
            for alt in alts[:3]:
                alt_name = alt.get("employee_name", "")
                alt_score = alt.get("score", "")
                alt_strs.append(f"{alt_name} ({alt_score})")
            while len(alt_strs) < 3:
                alt_strs.append("")

            iso_date = a.get("date", "")
            ws_review.append([
                _format_date_de(iso_date),
                _weekday_de(iso_date),
                a.get("start", ""),
                a.get("end", ""),
                a.get("shift_type", ""),
                a.get("working_area", ""),
                a.get("employee_name", ""),
                "Ja" if a.get("is_applicant") else "Nein",
                a.get("score", 0),
                reasoning,
                alt_strs[0],
                alt_strs[1],
                alt_strs[2],
            ])
        all_sheets.append(ws_review)

    # --- Eval Matrix sheet (if evaluation_matrix present) ---
    eval_matrix = plan.get("evaluation_matrix")
    if eval_matrix:
        ws_eval = wb.create_sheet("Eval Matrix")
        ws_eval.append(EVAL_MATRIX_COLS)

        # Build slot metadata lookup
        slot_meta: dict[str, dict] = {}
        for a in plan.get("assignments", []):
            slot_meta[a["slot_id"]] = {
                "date": a["date"], "start": a["start"],
                "end": a["end"], "area": a.get("working_area", ""),
            }
        for u in plan.get("unassigned", []):
            slot_meta[u["slot_id"]] = {
                "date": u["date"], "start": u["start"],
                "end": u["end"], "area": u.get("working_area", ""),
            }

        # Build selected lookup: slot_id -> employee_id
        selected_lookup: dict[str, str] = {}
        for a in plan.get("assignments", []):
            selected_lookup[a["slot_id"]] = a["ordio_employee_id"]

        for slot_id, evals in eval_matrix.items():
            meta = slot_meta.get(slot_id, {})
            selected_emp = selected_lookup.get(slot_id)
            for ev in evals:
                sd = ev.get("score_detail", {})
                ws_eval.append([
                    slot_id,
                    meta.get("date", ""),
                    meta.get("start", ""),
                    meta.get("end", ""),
                    meta.get("area", ""),
                    ev["employee_name"],
                    "Ja" if ev["blocked"] else "Nein",
                    pipe_join(ev.get("blocked_reasons", [])),
                    ev["score"],
                    *[sd.get(c, 0) for c in SCORE_COMPONENTS],
                    "Ja" if ev.get("is_applicant") else "Nein",
                    "TRUE" if (selected_emp and ev["employee_id"] == selected_emp) else "FALSE",
                ])
        all_sheets.append(ws_eval)

    # --- Directive Compliance sheet (if directives present in profile) ---
    if profile is not None:
        raw_directives = profile.get("_directives", [])
        if raw_directives:
            ws_compliance = wb.create_sheet("Directive Compliance")
            ws_compliance.append(DIRECTIVE_COMPLIANCE_COLS)
            compliance_rows = _build_directive_compliance(raw_directives, plan, profile)
            for row in compliance_rows:
                ws_compliance.append([row.get(c, "") for c in DIRECTIVE_COMPLIANCE_COLS])
            all_sheets.append(ws_compliance)

    # --- Golden Comparison sheet (if golden data provided) ---
    if golden_comparison and golden_comparison.get("per_slot"):
        from openpyxl.styles import PatternFill

        ws_golden = wb.create_sheet("Golden Comparison")
        golden_cols = [
            "slot_id", "date", "start", "end", "shift_type", "area",
            "golden_employee", "algo_employee", "match",
            "algo_score", "golden_in_alternatives", "golden_alt_rank",
        ]
        ws_golden.append(golden_cols)

        fill_map = {
            "exact_match": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "different_employee": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "algo_only": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "golden_only": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
        }

        for slot in golden_comparison["per_slot"]:
            ws_golden.append([
                slot.get("slot_id", ""),
                slot.get("date", ""),
                slot.get("start", ""),
                slot.get("end", ""),
                slot.get("shift_type", ""),
                slot.get("area", ""),
                slot.get("golden_employee", ""),
                slot.get("algo_employee", ""),
                slot.get("match", ""),
                slot.get("algo_score", ""),
                "Ja" if slot.get("golden_in_alternatives") else "Nein",
                slot.get("golden_alt_rank") or "",
            ])
            fill = fill_map.get(slot.get("match", ""))
            if fill:
                ws_golden.cell(row=ws_golden.max_row, column=9).fill = fill

        # Summary block
        gc = golden_comparison
        ws_golden.append([])
        ws_golden.append(["Summary"])
        ws_golden.append(["exact_match", gc.get("exact_match", 0)])
        ws_golden.append(["different_employee", gc.get("different_employee", 0)])
        ws_golden.append(["algo_only", gc.get("algo_only", 0)])
        ws_golden.append(["golden_only", gc.get("golden_only", 0)])
        ws_golden.append(["match_rate", f"{gc.get('match_rate', 0)}%"])

        # Per-employee sub-table
        ws_golden.append([])
        ws_golden.append(["Employee", "Algo Hours", "Golden Hours", "Delta"])
        for emp in gc.get("per_employee", []):
            ws_golden.append([
                emp.get("employee", ""),
                emp.get("algo_hours", 0),
                emp.get("golden_hours", 0),
                emp.get("delta", 0),
            ])
        all_sheets.append(ws_golden)

    _style_headers(all_sheets)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def _build_directive_compliance(
    raw_directives: list[dict],
    plan: dict[str, Any],
    profile: dict[str, Any],
) -> list[dict[str, Any]]:
    """Build directive compliance rows by checking plan outcomes against parsed rules."""
    from .directives import parse_directive

    # Build per-employee assignment totals
    emp_hours: dict[str, float] = {}
    emp_slots: dict[str, int] = {}
    emp_week_hours: dict[str, dict[str, float]] = {}  # name -> {week_key -> hours}
    emp_shift_types: dict[str, list[str]] = {}

    for a in plan.get("assignments", []):
        name = a.get("employee_name", "")
        hours = float(a.get("hours", 0))
        emp_hours[name] = emp_hours.get(name, 0) + hours
        emp_slots[name] = emp_slots.get(name, 0) + 1

        # Track weekly hours
        d = a.get("date", "")
        if d:
            try:
                from datetime import date as date_cls
                dt = date_cls.fromisoformat(d)
                monday = dt.fromordinal(dt.toordinal() - dt.weekday())
                wk = monday.isoformat()
            except (ValueError, TypeError):
                wk = ""
            if wk:
                emp_week_hours.setdefault(name, {})
                emp_week_hours[name][wk] = emp_week_hours[name].get(wk, 0) + hours

        # Track shift types
        st = a.get("shift_type", "")
        if st:
            emp_shift_types.setdefault(name, []).append(st.lower())

    rows: list[dict[str, Any]] = []
    for directive in raw_directives:
        d_id = directive.get("id", "")
        text = directive.get("text", "")
        employees_str = directive.get("employees", "")
        parsed = parse_directive(text)

        if not parsed:
            rows.append({
                "directive_id": d_id,
                "text": text,
                "employees": employees_str,
                "rule_field": "notes",
                "rule_value": text,
                "honored": "",
                "evidence": "Freitext -- manuelle Pruefung erforderlich",
            })
            continue

        # Get target employee names
        if employees_str.strip():
            names = [n.strip() for n in employees_str.split("|") if n.strip()]
        else:
            names = ["(global)"]

        for field, value in parsed.items():
            for name in names:
                honored = ""
                evidence = ""
                actual_hours = emp_hours.get(name, 0)
                actual_slots = emp_slots.get(name, 0)

                if field == "no_additional_shifts":
                    honored = "Ja" if actual_slots == 0 else "Nein"
                    evidence = f"{actual_slots} Zuweisungen"

                elif field == "max_monthly_hours":
                    honored = "Ja" if actual_hours <= float(value) else "Nein"
                    evidence = f"{actual_hours:.1f}h zugewiesen (Max: {value}h/Monat)"

                elif field == "max_weekly_hours" or field == "target_weekly_hours":
                    weeks = emp_week_hours.get(name, {})
                    if weeks:
                        max_week = max(weeks.values())
                        cap = float(value)
                        honored = "Ja" if max_week <= cap else "Nein"
                        evidence = f"Max Woche: {max_week:.1f}h (Limit: {cap}h/Woche)"
                    else:
                        honored = "Ja"
                        evidence = "0h zugewiesen"

                elif field == "max_additional_monthly_hours":
                    honored = "Ja" if actual_hours <= float(value) else "Nein"
                    evidence = f"{actual_hours:.1f}h zusaetzlich zugewiesen (Max: {value}h/Monat)"

                elif field == "preferred_shift_types":
                    types = emp_shift_types.get(name, [])
                    if not types:
                        honored = "Ja"
                        evidence = "Keine Zuweisungen"
                    else:
                        matching = sum(1 for t in types if any(v in t for v in value))
                        if matching == len(types):
                            honored = "Ja"
                        elif matching > 0:
                            honored = "Teilweise"
                        else:
                            honored = "Nein"
                        evidence = f"{matching}/{len(types)} Schichten passen"

                elif field == "preferred_working_areas":
                    honored = ""
                    evidence = "Bereichspraeferenz -- in Scoring beruecksichtigt"

                elif field == "notes":
                    honored = ""
                    evidence = "Freitext -- manuelle Pruefung erforderlich"

                rows.append({
                    "directive_id": d_id,
                    "text": text,
                    "employees": name,
                    "rule_field": field,
                    "rule_value": str(value),
                    "honored": honored,
                    "evidence": evidence,
                })

    return rows


def _build_unassigned_reasoning(u: dict[str, Any]) -> str:
    """Build German-language reasoning for an unassigned slot."""
    reason = u.get("reason", "")
    parts: list[str] = []

    if reason == "no_valid_candidate":
        parts.append("Keine validen Kandidaten.")
    elif reason == "all_candidates_blocked":
        parts.append("Alle Kandidaten blockiert.")
    elif reason:
        parts.append(f"Grund: {reason}.")

    top = u.get("top_candidates", [])
    if top:
        blocked_parts = []
        for c in top[:3]:
            name = c.get("employee_name", "")
            reasons = c.get("blocked_reasons", [])
            if reasons:
                blocked_parts.append(f"{name} ({', '.join(reasons)})")
        if blocked_parts:
            parts.append("Blockiert: " + "; ".join(blocked_parts) + ".")

    return " ".join(parts) if parts else "Nicht zugewiesen."


def _write_freigabe_sheet(ws, plan: dict[str, Any]) -> None:
    """Build the Freigabe (review/approval) sheet.

    Combines assigned + unassigned slots chronologically.  Read-only data
    columns are locked; decision columns (Entscheidung, Ersatz, Kommentar)
    are unlocked for reviewer input.  Includes data-validation dropdown,
    conditional fills, and a live summary block with COUNTIF formulas.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    # -- Colours ---------------------------------------------------------------
    HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    DATA_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    EDIT_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    OPEN_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    BLUE_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    SUMMARY_LABEL_FONT = Font(bold=True, name="Calibri", size=10)
    DATA_FONT = Font(name="Calibri", size=10)
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="CBD5E1"),
    )

    LOCKED = Protection(locked=True)
    UNLOCKED = Protection(locked=False)

    # -- Columns ---------------------------------------------------------------
    COLS = [
        ("Datum",        10),
        ("Tag",           5),
        ("Zeit",         13),
        ("Typ",          10),
        ("Bereich",      12),
        ("Zuweisung",    18),
        ("Bewerbung",     9),
        ("Score",         7),
        ("Begruendung",  60),
        ("Alternative 1", 22),
        ("Alternative 2", 22),
        ("Alternative 3", 22),
        ("Entscheidung", 14),
        ("Ersatz",       18),
        ("Kommentar",    40),
    ]
    DATA_COL_COUNT = 12   # A-L are read-only
    DECISION_COL = 13     # M = Entscheidung

    # -- Header row ------------------------------------------------------------
    for col_idx, (name, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.protection = LOCKED
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # -- Build merged + sorted row list ----------------------------------------
    rows: list[dict[str, Any]] = []

    for a in plan.get("assignments", []):
        reasoning = _build_reasoning(a)
        alts = a.get("alternatives", [])
        alt_strs = []
        for alt in alts[:3]:
            alt_name = alt.get("employee_name", "")
            alt_score = alt.get("score", "")
            alt_strs.append(f"{alt_name} ({alt_score})")
        while len(alt_strs) < 3:
            alt_strs.append("")

        rows.append({
            "date_iso": a.get("date", ""),
            "start": a.get("start", ""),
            "end": a.get("end", ""),
            "shift_type": a.get("shift_type", ""),
            "area": a.get("working_area", ""),
            "employee": a.get("employee_name", ""),
            "is_applicant": "Ja" if a.get("is_applicant") else "Nein",
            "score": a.get("score", 0),
            "reasoning": reasoning,
            "alt1": alt_strs[0],
            "alt2": alt_strs[1],
            "alt3": alt_strs[2],
            "is_open": False,
        })

    for u in plan.get("unassigned", []):
        reasoning = _build_unassigned_reasoning(u)
        rows.append({
            "date_iso": u.get("date", ""),
            "start": u.get("start", ""),
            "end": u.get("end", ""),
            "shift_type": u.get("shift_type", ""),
            "area": u.get("working_area", ""),
            "employee": "(offen)",
            "is_applicant": "-",
            "score": "-",
            "reasoning": reasoning,
            "alt1": "",
            "alt2": "",
            "alt3": "",
            "is_open": True,
        })

    rows.sort(key=lambda r: (r["date_iso"], r["start"]))

    # -- Write data rows -------------------------------------------------------
    for row_idx, r in enumerate(rows, 2):
        iso_d = r["date_iso"]
        values = [
            _format_date_de(iso_d),
            _weekday_de(iso_d),
            f"{r['start']}-{r['end']}",
            r["shift_type"],
            r["area"],
            r["employee"],
            r["is_applicant"],
            r["score"],
            r["reasoning"],
            r["alt1"],
            r["alt2"],
            r["alt3"],
            "",   # Entscheidung (empty, for reviewer)
            "",   # Ersatz
            "",   # Kommentar
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if col_idx <= DATA_COL_COUNT:
                cell.fill = DATA_FILL
                cell.protection = LOCKED
            else:
                cell.fill = EDIT_FILL
                cell.protection = UNLOCKED

            # Wrap text for reasoning column
            if col_idx == 9:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Highlight open slots
        if r["is_open"]:
            ws.cell(row=row_idx, column=6).fill = OPEN_FILL

    last_data_row = 1 + len(rows)

    # -- Data validation: Entscheidung dropdown --------------------------------
    if rows:
        dv = DataValidation(
            type="list",
            formula1='"Akzeptiert,Abgelehnt,Alternative,Manuell"',
            allow_blank=True,
        )
        dv.error = "Bitte waehle: Akzeptiert, Abgelehnt, Alternative oder Manuell"
        dv.errorTitle = "Ungueltige Eingabe"
        dv.prompt = "Entscheidung waehlen"
        dv.promptTitle = "Freigabe"
        col_letter = get_column_letter(DECISION_COL)
        dv.add(f"{col_letter}2:{col_letter}{last_data_row}")
        ws.add_data_validation(dv)

    # -- Conditional fill on Entscheidung column (manual via formula) -----------
    # openpyxl conditional formatting with formula rules
    from openpyxl.formatting.rule import CellIsRule

    col_letter = get_column_letter(DECISION_COL)
    cell_range = f"{col_letter}2:{col_letter}{last_data_row}"
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Akzeptiert"'], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Abgelehnt"'], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Alternative"'], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Manuell"'], fill=BLUE_FILL),
    )

    # -- Summary block below data ----------------------------------------------
    summary_row = last_data_row + 2
    metrics = plan.get("metrics", {})
    total = metrics.get("total_slots", len(rows))
    assigned = metrics.get("assigned_slots", sum(1 for r in rows if not r["is_open"]))
    unassigned = metrics.get("unassigned_slots", sum(1 for r in rows if r["is_open"]))
    fill_rate = metrics.get("fill_rate", round(assigned / total * 100, 1) if total else 0)

    summary_data = [
        ("Gesamt Schichten", total),
        ("Zugewiesen", assigned),
        ("Offen", unassigned),
        ("Fill Rate", f"{fill_rate}%"),
        ("Akzeptiert", f'=COUNTIF({col_letter}2:{col_letter}{last_data_row},"Akzeptiert")'),
        ("Abgelehnt", f'=COUNTIF({col_letter}2:{col_letter}{last_data_row},"Abgelehnt")'),
        ("Alternative/Manuell", f'=COUNTIF({col_letter}2:{col_letter}{last_data_row},"Alternative")'
                                f'+COUNTIF({col_letter}2:{col_letter}{last_data_row},"Manuell")'),
        ("Ohne Entscheidung", f'=COUNTBLANK({col_letter}2:{col_letter}{last_data_row})'),
    ]

    for i, (label, value) in enumerate(summary_data):
        r = summary_row + i
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = SUMMARY_LABEL_FONT
        label_cell.protection = LOCKED
        val_cell = ws.cell(row=r, column=2, value=value)
        val_cell.font = DATA_FONT
        val_cell.protection = LOCKED

    # -- Sheet protection (locks data columns, leaves M-O editable) ------------
    ws.protection.sheet = True
    ws.protection.password = ""  # no password -- structural protection only


def _write_guide_sheet(ws) -> None:
    """Populate the Guide sheet with static algorithm reference data."""
    _, Font, _ = _get_openpyxl()
    bold = Font(bold=True)

    # Section 1: Scoring Components
    ws.append(["SCORING COMPONENTS"])
    ws["A1"].font = bold
    ws.append(["Component", "Max Points", "Description"])
    scoring_rows = [
        ("applicant", 80, "Bonus for employees who applied for the shift"),
        ("rest", 40, "Remaining capacity toward monthly target hours"),
        ("fairness", 30, "Fewer hours this period = higher score"),
        ("role", 20, "Role matches shift type (exact=20, partial=10)"),
        ("preference", 15, "Employee preferences match shift (type, time, area)"),
        ("skill", 12, "Skills match working area"),
        ("fixed", 12, "Historical pattern -- employee regularly works this shift"),
        ("salary", -12, "Penalty when near salary cap"),
    ]
    for comp, pts, desc in scoring_rows:
        ws.append([comp, pts, desc])

    ws.append([])  # blank row

    # Section 2: Score Ranges
    ws.append(["SCORE RANGES"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Range", "Meaning"])
    ws.append([">= 100", "Strong recommendation (usually includes applicant bonus)"])
    ws.append([">= 80", "Solid recommendation"])
    ws.append(["< 80", "Weak recommendation, review alternatives"])

    ws.append([])

    # Section 3: Constraint Blocker Codes
    ws.append(["CONSTRAINT BLOCKER CODES"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Code", "Description"])
    constraint_rows = [
        ("overlap_same_day", "Already has a shift at the same time"),
        ("daily_hours_gt_10", "Would exceed 10h/day (ArbZG)"),
        ("weekly_hours_limit", "Would exceed weekly hours cap"),
        ("rest_lt_11h", "Less than 11h rest between shifts (ArbZG)"),
        ("consecutive_days_limit", "Would exceed max consecutive working days"),
        ("no_additional_shifts", "Profile rule: no additional shifts"),
        ("absence", "Employee is absent on this date"),
        ("already_has_shift_same_day", "Already assigned a shift this day"),
        ("monthly_hours_limit", "Would exceed monthly hours cap"),
        ("max_additional_monthly_hours", "Would exceed additional monthly hours cap"),
        ("max_weekly_hours", "Would exceed profile weekly hours cap"),
        ("max_salary_limit", "Would exceed monthly salary cap"),
    ]
    for code, desc in constraint_rows:
        ws.append([code, desc])

    ws.append([])

    # Section 4: Assignment Kind
    ws.append(["ASSIGNMENT KIND"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Kind", "Meaning"])
    ws.append(["applicant", "Employee applied and was assigned"])
    ws.append(["recommendation_without", "No applicants existed, algorithm recommended"])
    ws.append(["recommendation_despite", "Applicants existed but someone else scored higher"])


# ---------------------------------------------------------------------------
# Workbook 3: Comparison XLSX (multi-mechanism)
# ---------------------------------------------------------------------------

def render_comparison_xlsx(
    comparison: dict[str, Any],
    path: Path,
) -> Path:
    """Render a multi-mechanism comparison workbook.

    Sheets:
      1. Summary — Side-by-side KPIs per mechanism
      2. Assignments Comparison — Per-slot: mechanism assignments + agreement
      3. Fairness Comparison — Per-employee: hours/delta per mechanism
      4. Per-mechanism detail sheets
    """
    Workbook, Font, PatternFill = _get_openpyxl()

    wb = Workbook()
    all_sheets = []
    mechanisms = comparison.get("mechanisms", [])
    kpis = comparison.get("kpis", {})
    all_metrics = comparison.get("all_metrics", {})

    # Color fills for agreement
    fill_agree = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_majority = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_differ = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_unassigned = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    agreement_fills = {
        "all_agree": fill_agree,
        "majority": fill_majority,
        "all_differ": fill_differ,
        "all_unassigned": fill_unassigned,
    }

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["KPI"] + mechanisms)
    kpi_labels = {
        "fill_rate": "Fill Rate %",
        "assigned": "Assigned Slots",
        "total_slots": "Total Slots",
        "unassigned": "Unassigned Slots",
        "gini": "Gini Coefficient",
        "avg_score": "Average Score",
        "avg_abs_delta": "Avg |Delta| to Target (h)",
    }
    for kpi_key, label in kpi_labels.items():
        row = [label]
        for mech in mechanisms:
            val = kpis.get(mech, {}).get(kpi_key, "")
            if isinstance(val, float):
                val = round(val, 3)
            row.append(val)
        ws_summary.append(row)
    all_sheets.append(ws_summary)

    # --- Sheet 2: Assignments Comparison ---
    per_slot = comparison.get("per_slot", [])
    if per_slot:
        ws_assign = wb.create_sheet("Assignments Comparison")
        assign_cols = ["date"] + [f"{m}_employee" for m in mechanisms] + ["agreement"]
        ws_assign.append(assign_cols)
        for slot in per_slot:
            row = [slot.get("date", "")]
            for mech in mechanisms:
                row.append(slot.get(f"{mech}_employee", ""))
            agreement = slot.get("agreement", "")
            row.append(agreement)
            ws_assign.append(row)

            fill = agreement_fills.get(agreement)
            if fill:
                agree_col = len(assign_cols)
                ws_assign.cell(row=ws_assign.max_row, column=agree_col).fill = fill
        all_sheets.append(ws_assign)

    # --- Sheet 3: Fairness Comparison ---
    per_employee = comparison.get("per_employee", [])
    if per_employee:
        ws_fair = wb.create_sheet("Fairness Comparison")
        fair_cols = ["name", "target_hours"]
        for mech in mechanisms:
            fair_cols.extend([f"{mech}_hours", f"{mech}_delta"])
        ws_fair.append(fair_cols)
        for emp in per_employee:
            row = [emp.get("name", ""), emp.get("target_hours", "")]
            for mech in mechanisms:
                row.append(emp.get(f"{mech}_hours", ""))
                row.append(emp.get(f"{mech}_delta", ""))
            ws_fair.append(row)
        all_sheets.append(ws_fair)

    # --- Per-mechanism detail sheets ---
    for mech in mechanisms:
        m = all_metrics.get(mech, {})
        ws_detail = wb.create_sheet(f"{mech.capitalize()} Detail")
        ws_detail.append(["name", "assigned_hours", "assigned_slots", "target_hours", "delta"])
        for emp in m.get("fairness", {}).get("per_employee", []):
            ws_detail.append([
                emp.get("name", ""),
                emp.get("assigned_hours", 0),
                emp.get("assigned_slots", 0),
                emp.get("target_hours", ""),
                emp.get("delta", ""),
            ])
        all_sheets.append(ws_detail)

    _style_headers(all_sheets)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path
